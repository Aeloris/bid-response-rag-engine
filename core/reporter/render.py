# -*- coding: utf-8 -*-
"""报告渲染器：同一个 BidReport → Markdown / HTML / Excel（概念③ 单一产物源，多出口）。

- 业务内容只写在 build_report（service.py），这里全是"换个容器"，导出=换渲染器；
- HTML 一律 html.escape 所有模型/LLM 文本（防注入，哪怕 demo）；
- Markdown 表格单元格转义 | 与换行，避免破坏表结构。
"""
from __future__ import annotations

import html
from io import BytesIO

from .schemas import BidReport, ReportIssue, ReportPoint

# ------------------------------------------------------------ 文案映射
STATUS_LABEL = {
    "answered": "已应答",
    "answered_with_warn": "已应答 · 有WARN",
    "answered_with_block": "已应答 · 有BLOCK",
    "needs_material": "缺材料",
    "gap": "无上下文(gap)",
    "unanswered": "未应答",
}
VERDICT_LABEL = {
    "conform": "达标 CONFORM",
    "over": "正偏离 OVER ↑",
    "under": "负偏离 UNDER ↓",
    "unknown": "待人工 UNKNOWN ?",
}
VERDICT_HTML_CLS = {"conform": "ok", "over": "ok", "under": "bad", "unknown": "warn"}
SEV_LABEL = {"block": "BLOCK", "warn": "WARN", "info": "INFO"}


def _esc(text: object) -> str:
    return html.escape(str(text or ""), quote=False)


def _mdcell(text: object) -> str:
    """Markdown 表格单元格：去竖线换行，保留可读性。"""
    return str(text or "").replace("|", "\\|").replace("\r", " ").replace("\n", "<br>")


def _cit(c) -> str:
    parts = [f"[{c.ref}]"]
    if c.heading:
        parts.append(c.heading)
    if c.source:
        parts.append(f"（{c.source}）")
    return " ".join(parts)


# ------------------------------------------------------------ 一屏结论行（md/html 共用）
def _verdict_lines(report: BidReport) -> list[str]:
    v = report.verdict
    advise = "需拦截：存在废标级(BLOCK)风险，投出前必须人工复核" if v.escalation_required else "可投（无 BLOCK）"
    lines = [
        f"投出建议：{advise}",
        f"风险计数：BLOCK {v.block_count} · WARN {v.warn_count} · INFO {v.info_count}",
        f"应答覆盖：{v.answered_points}/{v.total_points} 已答 · ★ {v.star_answered}/{v.star_total}",
        f"数值核对：{v.calc_total} 条 → 达标 {v.calc_conform} · 正偏离 {v.calc_over} · 负偏离 {v.calc_under} · 待人工 {v.calc_unknown}",
    ]
    if v.needs_material:
        lines.append(f"待补材料：{'；'.join(v.needs_material)}")
    return lines


# ------------------------------------------------------------ Markdown
def render_markdown(report: BidReport) -> str:
    h, v = report.header, report.verdict
    out: list[str] = [f"# 应标应答包 · {h.tender_title or h.job_id}", ""]
    meta = [f"- 任务编号：`{h.job_id}`", f"- 报告生成：{h.generated_at}"]
    if h.source_file:
        meta.append(f"- 源文件：{h.source_file}")
    if h.buyer:
        meta.append(f"- 采购人：{h.buyer}")
    if h.deadline:
        meta.append(f"- 投标截止：{h.deadline}")
    out += meta + [""]

    if report.missing_artifacts:
        out += ["> ⚠️ 产物缺失：" + "、".join(report.missing_artifacts) + "（对应段落为空）", ""]

    out += ["## 一屏结论", ""]
    out += [f"- {line}" for line in _verdict_lines(report)] + [""]

    out += ["## 风险清单", ""]
    if not report.issues:
        out += ["- ✅ 未检出风险。", ""]
    else:
        out += ["| 严重级 | 评分点 | 类别 | 问题 | 出处 |", "|---|---|---|---|---|"]
        for i in report.issues:
            out.append(f"| {SEV_LABEL.get(i.severity.value, i.severity)} | `{_mdcell(i.point_id)}` | "
                       f"`{i.kind.value}` | {_mdcell(i.reason)} | {_mdcell(i.ref)} |")
        out += [""]

    if v.needs_material:
        out += ["## 待补材料", ""]
        out += [f"- {_mdcell(item)}" for item in v.needs_material] + [""]

    out += ["## 逐评分点应答", ""]
    for p in report.points:
        star = "★" if p.is_star else " "
        score = f" · {p.score}分" if p.score is not None else ""
        out += [f"### {p.point_id} [{STATUS_LABEL.get(p.status, p.status)}] {star}{score}", ""]
        out += [f"**应答要求**：{p.requirement}", ""]
        if p.answer.strip():
            out += ["**应答草稿**：", p.answer, ""]
        else:
            out += ["**应答草稿**：*（空——需人工或待补材料）*", ""]
        if p.citations:
            out += ["**引用**："] + [f"- {_mdcell(_cit(c))}" for c in p.citations] + [""]
        if p.missing_evidence:
            out += ["**待补**：" + "；".join(_mdcell(x) for x in p.missing_evidence), ""]
        if p.risks:
            out += ["**命中风险**：" + "；".join(f"{SEV_LABEL.get(r.severity.value, r.severity)} {r.kind.value}: {r.reason}"
                                                for r in p.risks), ""]
        if p.note:
            out += [f"*说明：{p.note}*", ""]
        out += ["---", ""]

    out += ["## 数值核对明细", ""]
    if not report.checks:
        out += ["- 无数值核对行（无数值要求或未跑核对）。", ""]
    else:
        out += ["| 参数 | ★ | 招标要求 | 我方声明 | 判定 | 说明 |", "|---|---|---|---|---|---|"]
        for c in report.checks:
            out.append(f"| {_mdcell(c.label)} | {'★' if c.star else ' '} | {_mdcell(c.requirement)} | "
                       f"{_mdcell(c.offer)} | {VERDICT_LABEL.get(c.verdict.value, c.verdict)} | {_mdcell(c.reason)} |")
        out += [""]

    out += ["## 附录", ""]
    if report.unparsed:
        out += ["**解析警告（未识别段落）**："] + [f"- {_mdcell(x)}" for x in report.unparsed] + [""]
    out += [f"- 生成于 {h.generated_at} · job `{h.job_id}` · 由 应标 Agent 自动装配，投出前请售前复核", ""]
    return "\n".join(out)


# ------------------------------------------------------------ HTML
_BASE_CSS = """
:root{--block:#b3261e;--warn:#b26a00;--ok:#1e6f3e;--ink:#1f2328;--mut:#656d76;--bg:#f6f8fa;--line:#d8dee4}
*{box-sizing:border-box}body{font-family:-apple-system,'Segoe UI','Microsoft YaHei',sans-serif;color:var(--ink);margin:0;background:#fff;line-height:1.55}
.wrap{max-width:960px;margin:0 auto;padding:24px 28px 64px}
h1{font-size:24px;border-bottom:2px solid var(--line);padding-bottom:10px}
h2{font-size:19px;margin-top:34px;border-left:4px solid #0969da;padding-left:10px}
h3{font-size:15px;margin:22px 0 6px}
.meta{color:var(--mut);font-size:13px;margin:6px 0 0}.meta code{background:var(--bg);padding:1px 5px;border-radius:4px}
.banner{border-radius:8px;padding:12px 16px;margin:16px 0;font-weight:600}
.banner.bad{background:#fdecea;color:var(--block);border:1px solid var(--block)}
.banner.ok{background:#e7f6ec;color:var(--ok);border:1px solid var(--ok)}
.banner.miss{background:#fff5e5;color:var(--warn);border:1px solid var(--warn)}
.grid{display:flex;gap:10px;flex-wrap:wrap;margin:12px 0}
.stat{flex:1 1 150px;background:var(--bg);border:1px solid var(--line);border-radius:8px;padding:10px 14px}
.stat b{display:block;font-size:22px}.stat span{color:var(--mut);font-size:12px}
table{border-collapse:collapse;width:100%;margin:10px 0;font-size:13px}
th,td{border:1px solid var(--line);padding:6px 9px;text-align:left;vertical-align:top}
th{background:var(--bg);font-weight:600;white-space:nowrap}
.ans{white-space:pre-wrap;background:var(--bg);border:1px solid var(--line);border-radius:6px;padding:10px 12px}
.chip{display:inline-block;font-size:12px;padding:1px 8px;border-radius:10px;border:1px solid var(--line);margin:0 4px 2px 0;color:var(--mut)}
.badge{display:inline-block;font-size:12px;padding:1px 8px;border-radius:10px;font-weight:600}
.bad{background:#fdecea;color:var(--block)}.warn{background:#fff5e5;color:var(--warn)}.ok{background:#e7f6ec;color:var(--ok)}.mut{background:var(--bg);color:var(--mut)}
.star{color:#cf222e}.req{color:var(--mut)}
"""


def _sev_cls(sev) -> str:
    return "bad" if sev == "block" else ("warn" if sev == "warn" else "mut")


def _verdict_cls(sev) -> str:
    return {"conform": "ok", "over": "ok", "under": "bad", "unknown": "warn"}.get(str(sev), "mut")


def _html_point_badge(p: ReportPoint) -> str:
    cls = {"answered": "ok", "answered_with_warn": "warn", "answered_with_block": "bad",
           "needs_material": "warn", "gap": "mut", "unanswered": "bad"}.get(p.status, "mut")
    return f'<span class="badge {cls}">{_esc(STATUS_LABEL.get(p.status, p.status))}</span>'


def _html_cits(p: ReportPoint) -> str:
    return "".join(f'<span class="chip">[{_esc(c.ref)}] {_esc(c.heading)} {_esc(c.source)}</span>' for c in p.citations)


def render_html(report: BidReport) -> str:
    h, v = report.header, report.verdict
    banner_cls = "bad" if v.escalation_required else "ok"
    banner_txt = ("⚠️ 存在废标级(BLOCK)风险 —— 投出前必须人工复核" if v.escalation_required
                  else "✅ 无 BLOCK —— 建议状态：可投（仍请售前终审）")
    miss = ""
    if report.missing_artifacts:
        miss = f'<div class="banner miss">⚠️ 产物缺失：{"、".join(_esc(x) for x in report.missing_artifacts)}（对应段落为空）</div>'

    stats = f"""
<div class="grid">
 <div class="stat"><b>{_esc(v.answered_points)}/{_esc(v.total_points)}</b><span>评分点已答</span></div>
 <div class="stat"><b>{_esc(v.star_answered)}/{_esc(v.star_total)}</b><span>★ 关键点已答</span></div>
 <div class="stat"><b>{_esc(v.block_count)}</b><span>BLOCK 风险</span></div>
 <div class="stat"><b>{_esc(v.warn_count)}</b><span>WARN 提示</span></div>
 <div class="stat"><b>{_esc(v.calc_under)}</b><span>数值负偏离</span></div>
</div>"""

    issues_html: list[str] = []
    if not report.issues:
        issues_html.append('<p>✅ 未检出风险。</p>')
    else:
        issues_html.append("<table><tr><th>严重级</th><th>评分点</th><th>类别</th><th>问题</th><th>出处</th></tr>")
        for i in report.issues:
            issues_html.append(
                f'<tr><td><span class="badge {_sev_cls(i.severity.value)}">{_esc(SEV_LABEL.get(i.severity.value, i.severity))}</span></td>'
                f'<td><code>{_esc(i.point_id)}</code></td><td><code>{_esc(i.kind.value)}</code></td>'
                f'<td>{_esc(i.reason)}</td><td class="req">{_esc(i.ref)}</td></tr>')
        issues_html.append("</table>")

    material_html = ""
    if v.needs_material:
        material_html = "<h2>待补材料</h2><ul>" + "".join(f"<li>{_esc(x)}</li>" for x in v.needs_material) + "</ul>"

    points_html: list[str] = []
    for p in report.points:
        star = '<span class="star">★</span>' if p.is_star else ""
        score = f' · {_esc(p.score)}分' if p.score is not None else ""
        req = p.requirement
        if p.answer.strip():
            ans = f'<div class="ans">{_esc(p.answer)}</div>'
        else:
            ans = '<div class="ans"><i>（空——需人工或待补材料）</i></div>'
        risk = ""
        if p.risks:
            risk = '<div class="req" style="margin-top:6px">命中风险：' + "；".join(
                f'<span class="badge {_sev_cls(r.severity.value)}">{_esc(SEV_LABEL.get(r.severity.value, r.severity))}</span> '
                f'<code>{_esc(r.kind.value)}</code> {_esc(r.reason)}' for r in p.risks) + "</div>"
        note = f'<div class="req">说明：{_esc(p.note)}</div>' if p.note else ""
        points_html.append(
            f'<h3 id="{_esc(p.point_id)}">{_esc(p.point_id)} {_html_point_badge(p)} {star}{_esc(score)}</h3>'
            f'<div class="req">应答要求：{_esc(req)}</div>'
            f'{ans}'
            f'<div style="margin-top:4px">{_html_cits(p)}</div>{risk}{note}')

    checks_html: list[str] = []
    if not report.checks:
        checks_html.append("<p>无数值核对行。</p>")
    else:
        checks_html.append("<table><tr><th>参数</th><th>★</th><th>招标要求</th><th>我方声明</th><th>判定</th><th>说明</th></tr>")
        for c in report.checks:
            checks_html.append(
                f'<tr><td>{_esc(c.label)}</td><td>{"★" if c.star else ""}</td><td>{_esc(c.requirement)}</td>'
                f'<td>{_esc(c.offer)}</td>'
                f'<td><span class="badge {_verdict_cls(c.verdict.value)}">{_esc(VERDICT_LABEL.get(c.verdict.value, c.verdict))}</span></td>'
                f'<td class="req">{_esc(c.reason)}</td></tr>')
        checks_html.append("</table>")

    unparsed = ""
    if report.unparsed:
        unparsed = "<h2>附录 · 解析警告</h2><ul>" + "".join(f"<li>{_esc(x)}</li>" for x in report.unparsed) + "</ul>"

    return f"""<!doctype html><html lang="zh"><head><meta charset="utf-8">
<title>应标应答包 · {_esc(h.tender_title or h.job_id)}</title><style>{_BASE_CSS}</style></head><body><div class="wrap">
<h1>应标应答包 · {_esc(h.tender_title or h.job_id)}</h1>
<div class="meta">job <code>{_esc(h.job_id)}</code> · 报告生成 {_esc(h.generated_at)}
{(' · 源文件 ' + _esc(h.source_file)) if h.source_file else ''}
{(' · 采购人 ' + _esc(h.buyer)) if h.buyer else ''}
{(' · 截止 ' + _esc(h.deadline)) if h.deadline else ''}</div>
<div class="banner {banner_cls}">{banner_txt}</div>{miss}
{stats}
<h2>一屏结论</h2><ul>{''.join(f'<li>{_esc(l)}</li>' for l in _verdict_lines(report))}</ul>
<h2>风险清单</h2>{''.join(issues_html)}
{material_html}
<h2>逐评分点应答</h2>{''.join(points_html)}
<h2>数值核对明细</h2>{''.join(checks_html)}
{unparsed}
<div class="meta" style="margin-top:40px">由 应标 Agent 自动装配 · 投出前请售前复核</div>
</div></body></html>"""


# ------------------------------------------------------------ 目录页 HTML（reports index）
def render_job_list_html(jobs: list[dict]) -> str:
    rows = []
    if not jobs:
        rows.append('<p>还没有任务。先跑一条：<code>curl -F "file=@fixtures/tender_sample.pdf" http://127.0.0.1:8000/tasks</code></p>')
    for j in jobs:
        esc = _esc
        badge_cls = "bad" if j.get("escalation") else "ok"
        badge_txt = "拦截" if j.get("escalation") else "可投"
        title = j.get("title") or j.get("job_id")
        rows.append(
            f'<tr><td><code>{esc(j.get("job_id"))}</code></td>'
            f'<td><a href="/reports/{esc(j.get("job_id"))}">{esc(title)}</a></td>'
            f'<td><span class="badge {esc(j.get("status"))}">{esc(j.get("status"))}</span></td>'
            f'<td><span class="badge {badge_cls}">{badge_txt}</span></td>'
            f'<td class="req">{esc(j.get("created_at"))}</td>'
            f'<td><a href="/reports/{esc(j.get("job_id"))}/export?fmt=md">md</a> · '
            f'<a href="/reports/{esc(j.get("job_id"))}/export?fmt=xlsx">xlsx</a></td></tr>')
    return f"""<!doctype html><html lang="zh"><head><meta charset="utf-8"><title>应标 Agent · 报告目录</title>
<style>{_BASE_CSS}</style></head><body><div class="wrap">
<h1>应标 Agent · 任务报告目录</h1>
<p class="meta">点击任务编号打开 HTML 报告；每行右侧可下载 md / xlsx。</p>
<table><tr><th>任务</th><th>标书</th><th>状态</th><th>建议</th><th>创建时间</th><th>导出</th></tr>
{''.join(rows)}</table></div></body></html>"""


# ------------------------------------------------------------ Excel
def render_xlsx_bytes(report: BidReport) -> bytes:
    """一个 .xlsx：概览 / 逐评分点应答 / 数值核对 / 风险清单 四 sheet。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    h, v = report.header, report.verdict
    wb = Workbook()
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBFF")
    thin = {"bad": "FFC7CE", "warn": "FFEB9C", "ok": "C6EFCE", "mut": "E7E6E6"}

    def _sheet(title: str):
        ws = wb.active if title == "概览" else wb.create_sheet()
        ws.title = title
        return ws

    def _hdr(ws, cols: list[str]):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
        for i, _ in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = 26

    def _colorize(ws, row: int, col: int, kind: str):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=thin.get(kind, "mut"))

    # 概览
    ws = _sheet("概览")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    meta = [
        ("任务编号", h.job_id), ("标书标题", h.tender_title), ("采购人", h.buyer or ""),
        ("投标截止", h.deadline or ""), ("源文件", h.source_file), ("报告生成", h.generated_at),
        ("投出建议", "需拦截（有 BLOCK）" if v.escalation_required else "可投（无 BLOCK）"),
        ("评分点应答", f"{v.answered_points}/{v.total_points}"), ("★ 关键点", f"{v.star_answered}/{v.star_total}"),
        ("BLOCK / WARN / INFO", f"{v.block_count} / {v.warn_count} / {v.info_count}"),
        ("数值核对", f"{v.calc_total} 条：达标{v.calc_conform} 正偏离{v.calc_over} 负偏离{v.calc_under} 待人工{v.calc_unknown}"),
        ("待补材料", "；".join(v.needs_material)),
    ]
    for k, val in meta:
        ws.append([k, val])
    ws.column_dimensions["A"].font = Font(bold=True)

    # 逐评分点应答
    ws = _sheet("应答包")
    _hdr(ws, ["编号", "状态", "★", "分值", "应答要求", "应答草稿", "引用", "待补材料", "命中风险", "说明"])
    for p in report.points:
        ws.append([p.point_id, STATUS_LABEL.get(p.status, p.status), "★" if p.is_star else "",
                   p.score or "", p.requirement, p.answer,
                   "；".join(_cit(c) for c in p.citations), "；".join(p.missing_evidence),
                   "；".join(f"{SEV_LABEL.get(r.severity.value, r.severity)} {r.kind.value}: {r.reason}" for r in p.risks),
                   p.note])
        _colorize(ws, ws.max_row, 2, "bad" if p.status in ("answered_with_block", "unanswered")
                  else ("warn" if p.status in ("answered_with_warn", "needs_material") else "ok"))

    # 数值核对
    ws = _sheet("数值核对")
    _hdr(ws, ["参数", "★", "招标要求", "我方声明", "判定", "待人工", "说明"])
    for c in report.checks:
        ws.append([c.label, "★" if c.star else "", c.requirement, c.offer,
                   VERDICT_LABEL.get(c.verdict.value, c.verdict), "是" if c.needs_human else "", c.reason])
        _colorize(ws, ws.max_row, 5, VERDICT_HTML_CLS.get(c.verdict.value, "mut"))

    # 风险清单
    ws = _sheet("风险清单")
    _hdr(ws, ["编号", "评分点", "严重级", "类别", "问题", "出处", "证据", "可自动修"])
    for i in report.issues:
        ws.append([i.id, i.point_id, SEV_LABEL.get(i.severity.value, i.severity), i.kind.value,
                   i.reason, i.ref, i.evidence, "是" if i.fixable else ""])
        _colorize(ws, ws.max_row, 3, i.severity.value if i.severity.value in thin else "mut")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
