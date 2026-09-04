# -*- coding: utf-8 -*-
"""Phase 7 报告器测试：纯 build/渲染 + 落盘→端点离线 e2e。

离线约定：不跑引擎 —— 纯测试手造强类型产物；e2e 用 JobStore 直接写 job 目录、
dependency_overrides 注入临时 store，再打 /reports* 端点（报告器只读落盘，快且确定）。
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient

from app.deps import get_job_store
from app.jobs import JobStore
from app.schemas import JobResult, JobStatus, PointBrief
from core.calculator.schemas import CalcSummary, NumericValue, OfferClaim, ParamCheck, ParamReq, Verdict
from core.generator.schemas import Citation, GenerationSummary, PointAnswer
from core.parser.schemas import ScorePoint, TenderDoc
from core.qa.schemas import IssueKind, IssueSeverity, QaIssue, QaReport
from core.reporter.schemas import BidReport
from core.reporter.service import build_report, build_report_from_artifacts
from core.reporter.render import (
    render_html,
    render_markdown,
    render_xlsx_bytes,
)

# ------------------------------------------------------------ 造样例产物
def _sample_doc() -> TenderDoc:
    return TenderDoc(
        tender_title="XX 市智慧园区智能化采购项目",
        buyer="XX 市机关事务管理局",
        deadline="2026-10-31 09:30",
        source_file="fixtures/tender_sample.pdf",
        score_points=[
            ScorePoint(id="SP-01", source="第三章", content="★ 质保期不少于 36 个月并提供原厂盖章承诺函",
                       score=5, is_star=True, evidence_type=["质保承诺函"]),
            ScorePoint(id="SP-02", content="提供近三年同类项目案例 3 份", score=4,
                       evidence_type=["项目案例"]),
        ],
    )


def _sample_answers() -> list[PointAnswer]:
    return [
        PointAnswer(point_id="SP-01", answer="我方提供质保 36 个月，并随标书附原厂盖章承诺函 [R1]",
                    citations=[Citation(ref="R1", chunk_id="c1", source="product-guide.md", heading="质保服务")]),
        PointAnswer(point_id="SP-02", answer="", missing_evidence=["同类项目案例"], note="检索无案例",
                    needs_human=True),
    ]


def _sample_checks() -> list[ParamCheck]:
    req = ParamReq(id="p1", label="质保期", topic="质保期", requirement="质保期不少于 36 个月",
                   numeric=NumericValue(value=36, unit="月", operator=">=", raw="不少于36个月"),
                   star=True, source="第三章")
    offer = OfferClaim(id="o1", label="质保期", topic="质保期", claim="质保 36 个月",
                       numeric=NumericValue(value=36, unit="月", operator="=", raw="36个月"),
                       source="product-guide.md")
    return [
        ParamCheck(req=req, offer=offer, verdict=Verdict.CONFORM, reason="36 ≥ 36，达标", needs_human=False),
        ParamCheck(req=ParamReq(id="p2", label="分辨率", topic="分辨率", requirement="≥ 400 万像素"),
                   offer=None, verdict=Verdict.UNKNOWN, reason="语料无对应能力，待人工", needs_human=True),
    ]


def _sample_result(escalation: bool = False) -> JobResult:
    issues = []
    if escalation:
        issues.append(QaIssue(id="q1", kind=IssueKind.STAR_UNDER, severity=IssueSeverity.BLOCK,
                              point_id="SP-01", reason="★ 质保期负偏离：答 12 月 < 要求 36 月",
                              ref="参数表", evidence="12个月"))
    issues.append(QaIssue(id="q2", kind=IssueKind.MATERIAL_GAP, severity=IssueSeverity.WARN,
                          point_id="SP-02", reason="缺同类项目案例，需售前补件", ref="SP-02"))
    qa = QaReport(
        issues=issues,
        block_count=1 if escalation else 0,
        warn_count=1,
        info_count=0,
        escalation_required=escalation,
        needs_material=["同类项目案例 x3"] if not escalation else ["质保承诺函（盖章）", "同类项目案例 x3"],
    )
    gen = GenerationSummary(total=2, answered=1, needs_material=["同类项目案例"], star_total=1, star_answered=1)
    return JobResult(
        job_id="abc123", status=JobStatus.DONE, tender_title="XX 市智慧园区智能化采购项目",
        score_points=[
            PointBrief(id="SP-01", score=5, is_star=True, content="★ 质保期不少于 36 个月…",
                       evidence_type=["质保承诺函"]),
            PointBrief(id="SP-02", score=4, content="提供近三年同类项目案例 3 份",
                       evidence_type=["项目案例"]),
        ],
        gen=gen, calc=CalcSummary(total=2, conform=1, over=0, under=0, unknown=1),
        qa=qa,
    )


# ------------------------------------------------------------ 纯 build_report
def test_build_report_structure_and_status() -> None:
    r = build_report(result=_sample_result(), doc=_sample_doc(),
                     answers=_sample_answers(), checks=_sample_checks())
    assert isinstance(r, BidReport)
    assert r.header.tender_title.startswith("XX 市")
    assert r.header.source_file == "fixtures/tender_sample.pdf"
    assert [p.point_id for p in r.points] == ["SP-01", "SP-02"]
    # 状态推导：SP-01 有正文无风险→answered；SP-02 空应答带 missing_evidence→needs_material
    assert r.points[0].status == "answered"
    assert r.points[1].status == "needs_material"
    assert r.points[0].citations[0].ref == "R1"
    # 核对明细展平 + 一屏结论
    assert len(r.checks) == 2
    assert r.checks[0].verdict == Verdict.CONFORM
    assert r.verdict.total_points == 2 and r.verdict.answered_points == 1
    assert r.verdict.calc_unknown == 1
    assert r.verdict.needs_material == ["同类项目案例 x3"]


def test_escalation_block_marks_point_and_banner() -> None:
    r = build_report(result=_sample_result(escalation=True), doc=_sample_doc(),
                     answers=_sample_answers(), checks=_sample_checks())
    assert r.verdict.escalation_required is True
    # BLOCK 排最前 + 命中点被标
    assert r.issues[0].severity == IssueSeverity.BLOCK
    assert r.points[0].status == "answered_with_block"
    assert any(p.severity == IssueSeverity.BLOCK for p in r.points[0].risks)
    html = render_html(r)
    assert "废标级" in html                       # 顶部红横幅
    assert "负偏离" in html                       # issue reason 进风险清单


def test_markdown_contains_sections_and_escape() -> None:
    result = _sample_result()
    # 注入脚本文本，验证 HTML 全 escape；Markdown 走 mdcell 换行转义
    answers = _sample_answers()
    answers[0].answer = "我方承诺 <script>alert(1)</script>\n 提供 36 月质保 [R1]"
    r = build_report(result=result, doc=_sample_doc(), answers=answers, checks=_sample_checks())
    md = render_markdown(r)
    assert "一屏结论" in md and "风险清单" in md and "待补材料" in md
    assert "应答要求" in md
    html = render_html(r)
    assert "<script>alert(1)</script>" not in html     # 已 escape
    assert "&lt;script&gt;" in html


def test_missing_artifacts_do_not_raise() -> None:
    r = build_report(result=_sample_result(), doc=None, answers=None, checks=None,
                     missing_artifacts=["04_calc_checks.json"])
    assert r.points[0].point_id == "SP-01"      # 用 result.score_points 兜底
    assert len(r.checks) == 0
    assert r.missing_artifacts == ["04_calc_checks.json"]


def test_xlsx_bytes_has_four_sheets() -> None:
    from openpyxl import load_workbook

    r = build_report(result=_sample_result(escalation=True), doc=_sample_doc(),
                     answers=_sample_answers(), checks=_sample_checks())
    data = render_xlsx_bytes(r)
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["概览", "应答包", "数值核对", "风险清单"]
    rows = list(wb["风险清单"].values)
    assert rows[0][0] == "编号"                  # 表头
    assert any("质保期" in str(cell) for row in rows[1:] for cell in row)


# ------------------------------------------------------------ L1：FAILED 任务不许渲染成可投
def _failed_result() -> JobResult:
    """qa 没跑（qa=None → block_count=0 / escalation=False），但引擎确实失败了的产物。"""
    return JobResult(
        job_id="abc123", status=JobStatus.FAILED, step="qa", error="qa 模型调用超时，任务中断",
        tender_title="XX 市智慧园区智能化采购项目",
        score_points=[
            PointBrief(id="SP-01", score=5, is_star=True, content="★ 质保期不少于 36 个月…",
                       evidence_type=["质保承诺函"]),
        ],
        gen=None, calc=None, qa=None,
    )


def test_failed_job_status_passthrough() -> None:
    r = build_report(result=_failed_result(), doc=None, answers=None, checks=None)
    assert r.status == "failed"
    assert r.step == "qa"
    assert "超时" in r.error
    # done 任务原样带 done（默认不透传成失败）
    assert build_report(result=_sample_result(), doc=_sample_doc(),
                        answers=_sample_answers(), checks=_sample_checks()).status == "done"


def test_failed_job_renderers_never_say_ke_tou() -> None:
    """回归：流水线 FAILED 但 qa 缺失 → 老代码 block_count=0/escalation=False 渲染成绿"可投"。
    现在必须三种导出都红标失败、不许出现"可投"。"""
    r = build_report(result=_failed_result(), doc=_sample_doc(), answers=_sample_answers(),
                     checks=_sample_checks())
    assert r.verdict.escalation_required is False  # 复现老 bug 的诱因
    assert r.verdict.block_count == 0

    md = render_markdown(r)
    assert "可投（无 BLOCK）" not in md and "引擎任务失败" in md and "qa 模型调用超时" in md

    html = render_html(r)
    assert "可投（无 BLOCK）" not in html and "引擎任务失败" in html and "qa 模型调用超时" in html

    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(render_xlsx_bytes(r)))
    overview = list(wb["概览"].values)
    joined = "\n".join(str(c) for row in overview for c in row)
    assert "可投（无 BLOCK）" not in joined and "失败——不可投" in joined and "qa 模型调用超时" in joined


def test_failed_job_index_page_shows_failed_badge(tmp_path: Path) -> None:
    """目录页回归：FAILED 任务没有 result（escalation=False）时列表曾假绿"可投"，须显示失败。"""
    store = JobStore(tmp_path / "jobs")
    job_id = "rep-failed"
    store.create(job_id, b"%PDF-1.4 demo")
    store.update(job_id, status=JobStatus.FAILED, step="qa", error="qa 超时")
    store.save_result(job_id, JobResult(job_id=job_id, status=JobStatus.FAILED, step="qa",
                                        error="qa 超时", tender_title="失败示例标"))
    from app.main import app

    app.dependency_overrides[get_job_store] = lambda: store
    try:
        with TestClient(app) as client:
            lst = client.get("/reports")
            assert lst.status_code == 200
            assert job_id in lst.text
            assert "失败" in lst.text
            assert "可投（无 BLOCK）" not in lst.text   # 唯一 FAILED 任务：不许出现绿色可投
            pg = client.get(f"/reports/{job_id}")
            assert pg.status_code == 200
            assert "引擎任务失败" in pg.text and "可投（无 BLOCK）" not in pg.text
    finally:
        app.dependency_overrides.pop(get_job_store, None)


# ------------------------------------------------------------ 落盘 → 端点 e2e（离线）
def _populate(store: JobStore) -> str:
    job_id = "rep-0001"
    store.create(job_id, b"%PDF-1.4 demo")
    store.update(job_id, status=JobStatus.DONE, step="done")
    store.save_step(job_id, "01_parse_doc", _sample_doc())
    store.save_step(job_id, "03_gen_answers", [a.model_dump(mode="json") for a in _sample_answers()])
    store.save_step(job_id, "04_calc_checks", [c.model_dump(mode="json") for c in _sample_checks()])
    store.save_result(job_id, _sample_result(escalation=True))
    return job_id


def test_report_endpoints_read_job_dir(tmp_path: Path) -> None:
    store = JobStore(tmp_path / "jobs")
    job_id = _populate(store)
    from app.main import app

    app.dependency_overrides[get_job_store] = lambda: store
    try:
        with TestClient(app) as client:
            # 目录页含该任务 + 拦截徽标
            lst = client.get("/reports")
            assert lst.status_code == 200
            assert job_id in lst.text and "拦截" in lst.text
            # HTML 报告页：标题/风险/逐点/拦截横幅
            pg = client.get(f"/reports/{job_id}")
            assert pg.status_code == 200
            assert "XX 市智慧园区" in pg.text
            assert "废标级" in pg.text and "风险清单" in pg.text
            assert "SP-01" in pg.text and "SP-02" in pg.text
            # md 下载
            md = client.get(f"/reports/{job_id}/export?fmt=md")
            assert md.status_code == 200
            assert "text/markdown" in md.headers["content-type"]
            assert "一屏结论" in md.text and "负偏离" in md.text
            # xlsx 下载可读
            xlsx = client.get(f"/reports/{job_id}/export?fmt=xlsx")
            assert xlsx.status_code == 200
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(xlsx.content))
            assert "风险清单" in wb.sheetnames
            # 未知任务 → 404
            assert client.get("/reports/ghost").status_code == 404
            assert client.get("/reports/ghost/export?fmt=md").status_code == 404
    finally:
        app.dependency_overrides.pop(get_job_store, None)


def test_failed_report_risk_section_not_claiming_clean() -> None:
    """F6 回归：FAILED 且无产物 → 风险清单不许显示"✅ 未检出风险"（半成品不能假装干净），
    应提示清单不可用。"""
    r = build_report(result=_failed_result(), doc=None, answers=None, checks=None)
    md = render_markdown(r)
    assert "风险清单不可用" in md and "未检出风险" not in md
    html = render_html(r)
    assert "风险清单不可用" in html and "未检出风险" not in html


def test_html_point_id_attribute_escaped() -> None:
    """F7 回归：point_id 进 id="{...}" 属性前必须转义引号，否则注入可逃出属性。"""
    from app.schemas import PointBrief
    from core.reporter.service import build_report

    result = _failed_result()
    result.score_points = [
        PointBrief(id='SP-01" onmouseover="alert(1)', score=5, is_star=True,
                   content="★ 质保期不少于 36 个月", evidence_type=["质保承诺函"]),
    ]
    r = build_report(result=result, doc=None, answers=None, checks=None)
    html = render_html(r)
    assert 'onmouseover="alert(1)' not in html      # 引号已被转义，payload 不可能原样落地
    assert "&quot;" in html


def test_running_job_index_badge_not_ke_tou(tmp_path: Path) -> None:
    """F5 回归：执行中(pending/running)任务在目录页显示"进行中"，不许假绿"可投"。"""
    from app.main import app

    store = JobStore(tmp_path / "jobs")
    job_id = "rep-running"
    store.create(job_id, b"%PDF-1.4 demo")
    store.update(job_id, status=JobStatus.RUNNING, step="generate")

    app.dependency_overrides[get_job_store] = lambda: store
    try:
        with TestClient(app) as client:
            lst = client.get("/reports")
            assert lst.status_code == 200
            assert job_id in lst.text
            assert "进行中" in lst.text
            assert "可投" not in lst.text            # 唯一任务在跑 → 全文不得出现"可投"
    finally:
        app.dependency_overrides.pop(get_job_store, None)
